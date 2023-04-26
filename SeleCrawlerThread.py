# 浏览器操作类
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options as options

# 正则类
import re

import os

import time

# Excel文件转换类
import xlwings as xw

# xlsx文件写入并保存
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# xls文件写入并保存
from xlutils.copy import copy
import xlrd
import xlwt

# 子窗口类
from MassageWindow import MassageWindow

# PyQt5类
from PyQt5.QtCore import QThread,pyqtSignal

"""
线程类：
用于爬取数据和Excel文件写入
"""
class seleCrawlerThread(QThread):
    FinishSignal = pyqtSignal(bool)
    WarningSignal = pyqtSignal(int)
    def __init__(self) -> None:
        super(seleCrawlerThread, self).__init__()
        # 实例化对象
        self.m_MassageWindow = None                 # 实例化爬取信息类(主要记录那些文件中的编号没有数据)

        # 成员变量
        self.loginUrl:str = ""  # 登录界面URL
        self.searchUrl:str = "http://202.117.114.3/chedeng/Product/ProductSeek_zyx.asp"  # 数据爬取界面URL
        self.driver_path:str = r'./geckodriver.exe' # 代理软件路径
        self.binary_path:str = 'C:\\Users\\yueshenl\\AppData\\Local\\Mozilla Firefox\\firefox.exe'  # 浏览器路径
        self.Username:str = 'hascovision'   # 用户名
        self.Password:str = 'hascovision'   # 密码
        self.PDMNum:str = ""                # 总成编号
        self.PDMName:str = ""               # 主机厂代号
        self.PDMNumber:str = ""             # 灯具总成号
        self.FilePath:str = ""              # 表格文件路径
        self.FilePathList:list = []         # 表格文件地址列表
        self.ZCName:str = ""                # 所属灯具总成名称
        self.ZCNumber:str = ""              # 所属灯具总成编号
        self.LightNum:int = 0               # 灯具数量
        self.XlsNum:int = 0                 # Xls文件的数量
        self.NoInformationFileName = ""     # 没有数据的文件名
        self.WrangMassageDic:dict = {}      # 错误信息字典
        self.IsFolder:bool = False          # 是否选择文件夹

        # 网页操作变量
        self.options = None
        self.serv = None
        self.browser = None

        # Excel文件操作变量
        self.WorkBookRD = None              # xlrd 文件打开类
        self.WorkBookWT = None              # xlwt 文件写入类
        self.WorkBookOP = None              # openpyxl 文件打开写入类
        self.PDMIndexList:list = []         # PDM号在表格中的列位置
        self.SheetList:list = []            # Sheet列表
        self.IndexList:list = []            # 索引值存储列表(列表长度为sheet的数量)
        self.TitleDict:dict = {}            # 存储表头行列信息的字典
        self.PdmAllDict:dict = {}           # 字典:key:sheet名称 value:PDM号列表
        self.IndexDict:dict = {}            # 字典:key:sheet名称 value:PDM号行号列表
        self.TitleIndexList:int = 0         # 表头所在的行数
        self.NewMassageIndex:int = 0        # 填入数据的列位置  (n,n+1,n+2)
        self.ITEMRow:int= 0                 # 记录item表头的行数
        self.ITEMCol:int = 0                # 记录item表头的列数
        self.ZJCRow:int = 0                 # 记录主机厂表头的行数
        self.ZJCCol:int = 0                 # 记录主机厂表头的列数
        self.IsFind:bool = False            # 记录是否查找到数据
        # 用于写入Excel表(之前的逻辑有问题改成按文件写入)
        self.ZCNameDict:dict = {}           # 接收所有的主机厂信息(空信息赋值为"")
        self.ZCNumberDict:dict = {}         # 接收所有的灯具号(空信息赋值为"")
        self.LightNumDict:dict = {}         # 接收所有的灯具数量

        self.IsSpecial:int = 0              # 是否为特殊情况(目前有两种1.表格横向排列、2.只读总成编号)

    """
    获取登录界面URL，并新建操作对象
    * 方法需要修改,发现打开一个模拟窗口进行多次操作数据会爬不到(目前不清楚原因)
    """
    def SetLoginUrl(self,iLoginUrl:str,iBinary:str):
        self.loginUrl = iLoginUrl                                           # 设置登录界面URL
        # self.driver_path = iDriver
        self.binary_path = iBinary
        self.options = webdriver.FirefoxOptions()                           # 设置代理种类(火狐浏览器)
        # 设置打开代理形式
        self.options.add_argument('--headless')                         
        self.options.add_argument('--disable-gpu')
        self.options.binary_location = self.binary_path                     # 设置浏览器执行文件路径
        self.serv = Service(self.driver_path)                                    # 设置代理执行文件路径
        self.browser = webdriver.Firefox(service=self.serv,options=self.options) # 启动代理
        self.browser.get(self.loginUrl)                                     # 获取登录界面URL,用于发送请求
        userNameTag = self.browser.find_elements(By.ID, 'UserName')         # 查找元素->用户名 未查找说明请求访问失败
        if (len(userNameTag)) <= 0:
            self.WarningSignal.emit(1)                                      # 登录界面URL失效
            self.browser.quit()
            return
        self.browser.find_elements(By.ID, 'UserName')[0].send_keys(self.Username)    # 找到账号框并输入账号
        self.browser.find_elements(By.ID, 'UserPasswd')[0].send_keys(self.Password)  # 找到密码框并输入密码
        self.browser.find_elements(By.NAME, 'OK')[0].click()                         # 找到登陆按钮并点击
        self.browser.get(self.searchUrl)                                             # 获取数据界面URL用于网页请求

    """
    获取单文件路径
    """
    def SetFilePath(self,iFP:str):
        self.FilePath = iFP
        print("self.FilePath = ",self.FilePath)
    
    """
    获取文件列表
    """
    def SetFilePathList(self,iFPList:list):
        self.FilePathList = iFPList                        # 获取Excel文件路径列表
    
    """
    获取PDM编号(item编号)，并模拟点击查询按钮
    """
    def SetPDMNumber(self,iPDM:str):
        self.IsNone = False                                # 初始化默认不为空
        PDMpattern=re.compile("@@")
        PDMresult = PDMpattern.split(iPDM)
        self.PDMNum = PDMresult[0]                                 # 获取PDM编码
        print("读取的PDM号为:  ",self.PDMNum)
        PDMLIST = re.findall(r'[A-Z]\d{8}', str(self.PDMNum), flags=re.IGNORECASE)  # 有的表格一个单元格中PDM号有多个
        # 数据初始化
        self.ZCName = ""
        self.ZCNumber = ""
        self.LightNum = 0
        for i in PDMLIST:
            try:
                # if len(i) != 9:
                #     return 
                self.browser.find_elements(By.NAME, 'ItemID')[0].clear()              # 清空填入的数据
                self.browser.find_elements(By.NAME, 'ItemID')[0].send_keys(i)         # 输入PDM编码
                self.browser.find_elements(By.CSS_SELECTOR, 'input[type="button"]')[0].click()  # 点击查询
                resTdList = self.browser.find_elements(By.CSS_SELECTOR, "table[id='resulttable'] > tbody > tr > td")  # 查询结果行数据
                if (len(resTdList[5].text) == 0):  
                    self.IsNone = True
                    ZCName = "无"
                    ZCNumber = "无"
                    self.ZCName += ZCName               # 信息拼接
                    self.ZCNumber += ZCNumber           # 信息拼接
                else:
                    # 将括号替换为\n方便在表格中显示
                    ZCName = resTdList[5].text.replace("[","").replace("]","\n")
                    ZCNumber = resTdList[6].text.replace("[","").replace("]","\n")
                    self.ZCName += ZCName               # 信息拼接
                    self.ZCNumber += ZCNumber           # 信息拼接
                    # 获取灯具的数量
                    pattern=re.compile(r"[[](.*?)[]]")
                    self.LightNum += len(pattern.findall(resTdList[5].text))
            except:
                print("异常了 = ",iPDM)
                self.IsNone = True
                ZCName = "异常请手动查询"
                ZCNumber = "异常请手动查询"
                self.ZCName += ZCName               # 信息拼接
                self.ZCNumber += ZCNumber           # 信息拼接
        self.ZCNameDict[iPDM] = self.ZCName
        self.ZCNumberDict[iPDM] = self.ZCNumber
        self.LightNumDict[iPDM] = self.LightNum
        print("self.ZCName = ",self.ZCName)
        print("self.ZCNumber = ",self.ZCNumber)
        print("self.LightNum = ",self.LightNum)

    """
    获取Xls文件的个数
    """
    def SetXlsNum(self,iNum:int):
        self.XlsNum = iNum                                       # 获取Xls文件数量

    """
    打开表格文件并获取相应的信息
    写入数据会更改表格格式方法暂存(不启用)
    """
    # def ExcelRead_Xls(self,iFilePath:str):
    #     """
    #     提取检索
    #     注意: 
    #     1.xlrd 读取Excel文件时，表格为空时，读取内容为字符串 ""
    #     2.可能有忘记写标题的情况在最大列数后加三列
    #     3.写入数据索引开头为0
    #     4.只保存第一次找到关键字的那一行,可能出现一行出现多个关键字，记录多个列值只记录一个行值
    #     """
    #     self.WorkBookRD = xlrd.open_workbook(iFilePath)          # 打开Excel文件
    #     self.SheetList = self.WorkBookRD.sheet_names()           # 获取所有表格

    #     # 需要通过判断去掉没有数据的Sheet
    #     for i in self.SheetList:                                 # 按每个sheet读取数据
    #         sheet = self.WorkBookRD.sheet_by_name(i)             # 读取Sheet
    #         # 初始化
    #         self.ITEMRow = 0
    #         self.ITEMCol = 0
    #         self.ZJCRow = 0
    #         self.ZJCCol = 0
    #         RCList = []
    #         self.IsFind = False
    #         for j in range(sheet.nrows):                         # 遍历行数
    #             for z in range(sheet.ncols):                     # 遍历相应的列
    #                 # 如果为空跳过
    #                 if sheet.cell_value(j, z) == "":   
    #                     continue
    #                 # 如果在不区分大小写的情况下找到‘item’关键字则记录行数和列数
    #                 if len(re.findall('item', str(sheet.cell_value(j, z)), flags=re.IGNORECASE)) > 0 and len(str(sheet.cell_value(j, z)))<50:
    #                     self.ITEMRow = j
    #                     self.ITEMCol = z
    #                     self.ZJCRow = j         # 因为主机厂的表头与'item'在同一行，行数一致
    #                     self.IsFind = True
    #                     break                   # 跳出循环
    #             if self.IsFind:  # 如果找到了第一组值则跳出整体的行循环(因为都在同一列)
    #                 break
    #         # 查找'主机厂'的表头，遍历所有列
    #         for k in range(sheet.ncols):
    #             if sheet.cell_value(self.ITEMRow, k) == "主机厂": # 如果查找到主机厂则记录所在的列，其后两列直接运算可得到
    #                 self.ZJCCol = k
    #         if self.ZJCCol == 0:                              # 防止出现用户忘记写表头自行添加
    #             self.ZJCCol = sheet.ncols+1
    #         # 表头行列信息存储
    #         RCList.append(self.ITEMRow)
    #         RCList.append(self.ITEMCol)
    #         RCList.append(self.ZJCRow)
    #         RCList.append(self.ZJCCol)
    #         print("RCList = ",RCList)
    #         # 压入字典
    #         self.TitleDict[i] = RCList
    #     # pdm号与行索引值提取
    #     for i in self.SheetList:
    #         sheet = self.WorkBookRD.sheet_by_name(i)
    #         PdmAllList = []
    #         PdmIndexList = {}
    #         for j in range(self.TitleDict[i][0]+1,sheet.nrows): # 取表头的下一行
    #             if len(re.findall(r'[A-Z]\d{8}', str(sheet.cell_value(j, self.TitleDict[i][1])), flags=re.IGNORECASE)) > 0:
    #                 PdmAllList.append(sheet.cell_value(j, self.TitleDict[i][1]))
    #                 PdmIndexList[sheet.cell_value(j, self.TitleDict[i][1])] = j
    #         self.PdmAllDict[i] = PdmAllList
    #         self.IndexDict[i] = PdmIndexList
    #     print("self.TitleDict = ",self.TitleDict)
    #     print("self.PdmAllDict = ",self.PdmAllDict)
    #     print("self.IndexDict = ",self.IndexDict)

    def ExcelRead_Xlsx(self,iFilePath:str):
        """
        提取检索
        注意: openpyxl 读取Excel文件时，表格为空时，读取内容为 None
        2.可能有忘记写标题的情况在最大列数后加三列
        3.写入数据索引开头为1
        4.只保存第一次找到关键字的那一行,可能出现一行出现多个关键字，记录多个列值只记录一个行值
        """
        print("iFilePath = ",iFilePath)
        self.WorkBookOP = load_workbook(iFilePath)               # 打开Excel文件
        self.SheetList = self.WorkBookOP.get_sheet_names()       # 获取所有表格

        # 需要通过判断去掉没有数据的Sheet
        for i in self.SheetList:
            sheet = self.WorkBookOP[i]                           # 读取sheet
            # 初始化
            self.ITEMRow = 1
            self.ITEMCol = 1
            self.ZJCRow = 1
            self.ZJCCol = 1
            RCList = []                                              # 临时变量存储行列信息
            self.IsFind = False
            if sheet.max_row == 1:
                RCList.append(1)
                RCList.append(1)
                RCList.append(1)
                RCList.append(1)
                self.TitleDict[i] = RCList
                continue
            for j in range(sheet.min_row,sheet.max_row+1):             # 遍历行数
                for k in range(sheet.min_column,sheet.max_column):   # 遍历列数
                    # 如果为空跳过
                    if sheet.cell(j, k).value == None:
                        continue
                    # 如果在不区分大小写的情况下找到‘item’关键字则记录行数和列数
                    if len(re.findall('item',str(sheet.cell(j, k).value),flags=re.IGNORECASE))>0 and len(str(sheet.cell(j, k).value))<50:
                        print(i," = ",j," 行 ",k," 列 ",sheet.cell(j, k).value)
                        print("第 ",j, " 行 ", "第 ",k," 列")
                        self.ITEMRow = j
                        self.ITEMCol = k
                        self.ZJCRow = j                              # 因为主机厂的表头与'item'在同一行，行数一致
                        self.IsFind = True
                        break                                        # 跳出循环
                if self.IsFind:                # 如果找到了第一组值则跳出整体的行循环(因为都在同一列)
                    break
            # 查找'主机厂'的表头，遍历所有列
            for z in range(sheet.min_column,sheet.max_column):
                if sheet.cell(self.ITEMRow, z).value == "主机厂":    # 如果查找到主机厂则记录所在的列，其后两列直接运算可得到
                    self.ZJCCol = z
            if self.ZJCCol == 1:                                # 防止出现用户忘记写表头自行添加
                self.ZJCCol = sheet.max_column+1
            RCList.append(self.ITEMRow)
            RCList.append(self.ITEMCol)
            RCList.append(self.ZJCRow)
            RCList.append(self.ZJCCol)
            self.TitleDict[i] = RCList
        for i in self.SheetList:
            sheet = self.WorkBookOP[i]
            PdmAllList = []
            PdmIndexList = {}
            for j in range(self.TitleDict[i][0],sheet.max_row+1):
                if len(re.findall(r'[A-Z]\d+',str(sheet.cell(j, self.TitleDict[i][1]).value),flags=re.IGNORECASE))>0:
                    PdmAllList.append(sheet.cell(j, self.TitleDict[i][1]).value+"@@"+str(j))
                    # 因为有同号的情况,修改字典的key值
                    PdmIndexList[sheet.cell(j, self.TitleDict[i][1]).value+"@@"+str(j)] = j
            self.PdmAllDict[i] = PdmAllList
            self.IndexDict[i] = PdmIndexList

    """
    特殊情况一：有横向表格
    """
    def HorizontalArrangementExcel(self,iFilePath:str):
        self.WorkBookOP = load_workbook(iFilePath)               # 打开Excel文件
        self.SheetList = self.WorkBookOP.get_sheet_names()       # 获取所有表格

        # 需要通过判断去掉没有数据的Sheet
        for i in self.SheetList:
            sheet = self.WorkBookOP[i]                           # 读取sheet
            # 初始化
            self.ITEMRow = 1
            self.ITEMCol = 1
            self.ZJCRow = 1
            self.ZJCCol = 1
            RCList = []                                              # 临时变量存储行列信息
            self.IsFind = False
            if sheet.max_row == 1:
                RCList.append(1)
                RCList.append(1)
                RCList.append(1)
                RCList.append(1)
                self.TitleDict[i] = RCList
                continue
            for j in range(sheet.min_row,sheet.max_row):
                for k in range(sheet.min_column,sheet.max_column):
                    if sheet.cell(j, k).value == None:
                        continue
                    if len(re.findall('ITEM No',str(sheet.cell(j, k).value)))>0 or sheet.cell(j, k).value == "主机厂":
                        print(i," = ",j," 行 ",k," 列 ",sheet.cell(j, k).value+"@@"+str(j))
                        print("第 ",j, " 行 ", "第 ",k," 列")
                        self.ITEMRow = j
                        self.ITEMCol = k
                        RCList.append(self.ITEMRow)
                        RCList.append(self.ITEMCol)
                        self.IsFind = True
                if self.IsFind:
                    break
            self.TitleDict[i] = RCList
        # 获取Item号
        for i in self.SheetList:
            sheet = self.WorkBookOP[i]
            PdmAllList = []
            PdmIndexList = {}
            for z in range(int(len(self.TitleDict[i])/4)):
                for j in range(self.TitleDict[i][0],sheet.max_row+1):
                    if len(re.findall(r'[A-Z]\d+',str(sheet.cell(j, self.TitleDict[i][z*4+1]).value),flags=re.IGNORECASE))>0:
                        PdmAllList.append(sheet.cell(j, self.TitleDict[i][z*4+1]).value+"@@"+str(self.TitleDict[i][z*4+3])+"@@"+str(j))
                        # 因为有同号的情况,修改字典的key值
                        PdmIndexList[sheet.cell(j, self.TitleDict[i][z*4+1]).value+"@@"+str(self.TitleDict[i][z*4+3])+"@@"+str(j)] = j
            self.PdmAllDict[i] = PdmAllList
            self.IndexDict[i] = PdmIndexList

    """
    特殊情况二：只识别总成编号
    """
    def AssemblyItemExcel(self,iFilePath:str):
        self.WorkBookOP = load_workbook(iFilePath)               # 打开Excel文件
        self.SheetList = self.WorkBookOP.get_sheet_names()       # 获取所有表格

        # 需要通过判断去掉没有数据的Sheet
        for i in self.SheetList:
            sheet = self.WorkBookOP[i]                           # 读取sheet
            # 初始化
            self.ITEMRow = 1
            self.ITEMCol = 1
            self.ZJCRow = 1
            self.ZJCCol = 1
            RCList = []                                              # 临时变量存储行列信息
            self.IsFind = False
            if sheet.max_row == 1:
                RCList.append(1)
                RCList.append(1)
                RCList.append(1)
                RCList.append(1)
                self.TitleDict[i] = RCList
                continue
            for j in range(sheet.min_row,sheet.max_row):             # 遍历行数
                for k in range(sheet.min_column,sheet.max_column):   # 遍历列数
                    # 如果为空跳过
                    if sheet.cell(j, k).value == None:
                        continue
                    # 如果在不区分大小写的情况下找到‘item’关键字则记录行数和列数
                    if len(re.findall('总成',str(sheet.cell(j, k).value),flags=re.IGNORECASE))>0 and len(str(sheet.cell(j, k).value))<50:
                        print(i," = ",j," 行 ",k," 列 ",sheet.cell(j, k).value)
                        print("第 ",j, " 行 ", "第 ",k," 列")
                        self.ITEMRow = j
                        self.ITEMCol = k
                        self.ZJCRow = j                              # 因为主机厂的表头与'item'在同一行，行数一致
                        break                                        # 跳出循环
                if self.ITEMRow != 1 or self.ITEMCol != 1:
                    break 
            # 查找'主机厂'的表头，遍历所有列
            for z in range(sheet.min_column,sheet.max_column):
                if sheet.cell(self.ITEMRow, z).value == "主机厂":    # 如果查找到主机厂则记录所在的列，其后两列直接运算可得到
                    self.ZJCCol = z
            if self.ZJCCol == 0:                                # 防止出现用户忘记写表头自行添加
                self.ZJCCol = sheet.max_column+1
            RCList.append(self.ITEMRow)
            RCList.append(self.ITEMCol)
            RCList.append(self.ZJCRow)
            RCList.append(self.ZJCCol)
            self.TitleDict[i] = RCList
        # 获取Item号
        for i in self.SheetList:
            sheet = self.WorkBookOP[i]
            PdmAllList = []
            PdmIndexList = {}
            for j in range(self.TitleDict[i][0],sheet.max_row+1):
                if len(re.findall(r'[A-Z]\d+',str(sheet.cell(j, self.TitleDict[i][1]).value),flags=re.IGNORECASE))>0:
                    PdmAllList.append(sheet.cell(j, self.TitleDict[i][1]).value+"@@"+str(j))
                    # 因为有同号的情况,修改字典的key值
                    PdmIndexList[sheet.cell(j, self.TitleDict[i][1]).value+"@@"+str(j)] = j
            self.PdmAllDict[i] = PdmAllList
            self.IndexDict[i] = PdmIndexList
            

    """
    将相应的信息写入表格文件
    输入:
    文件存储地址、Excel文件对象、表头字典、PDM号字典、PDM行索引字典、Sheet列表
    字典详解：
    1.表头字典------->key:sheet名 value:对应表头的行列值列表(ITEM号所在行和列、主机厂所在行和列)
    2.PDM号字典------>key:sheet名 value:该sheet里的PDM号列表
    3.PDM行索引字典-->key:sheet名 value:数据字典-->key:PDM号 value:行索引
    方法暂存(不启用)
    """
    # def ExcelWrite_Xls(self,iFilePath:str,iExcelClass,iTitleDict:dict,iPdmAllDict:dict,iIndexDict:dict
    #                    ,iSheetList:list,iAllZCName:dict,iAllZCNumber:dict,iAllLightNum:dict):
    #     """
    #     引用实例化的表格类来写数据
    #     """
    #     print("开始写数据！！！！！！！")
    #     print("iTitleDict = ",iTitleDict)
    #     self.WorkBookWT = copy(iExcelClass)                      # 获取Excel文件变量
    #     for i in range(len(iSheetList)):
    #         # xlwt 读取sheet有点不同
    #         sheet = self.WorkBookWT.get_sheet(i)             # 读取Sheet
    #         if iPdmAllDict[iSheetList[i]] != []:
    #             for j in iPdmAllDict[iSheetList[i]]:
    #                 # 设置单元格对齐方式
    #                 alignment = xlwt.Alignment()
    #                 # 0x01(左端对齐)、0x02(水平方向上居中对齐)、0x03(右端对齐)
    #                 alignment.horz = 0x02
    #                 # 0x00(上端对齐)、 0x01(垂直方向上居中对齐)、0x02(底端对齐)
    #                 alignment.vert = 0x01
    #                 # 设置自动换行
    #                 alignment.wrap = 1
    #                 # 格式设定
    #                 style = xlwt.XFStyle()
    #                 style.alignment = alignment
    #                 # 数据写入 style格式设定
    #                 sheet.write(iIndexDict[iSheetList[i]][j],iTitleDict[iSheetList[i]][3],iAllZCName[iSheetList[i]][j],style)                     # 主机厂
    #                 sheet.write(iIndexDict[iSheetList[i]][j],iTitleDict[iSheetList[i]][3]+1,iAllZCNumber[iSheetList[i]][j],style)                 # 灯具总成号
    #                 sheet.write(iIndexDict[iSheetList[i]][j],iTitleDict[iSheetList[i]][3]+2,iAllLightNum[iSheetList[i]][j],style)                 # 借用灯具数量
    #     print("结束！！！！！！！")
    #     # 文件保存
    #     print("保存！！！！！！！")
    #     self.WorkBookWT.save(iFilePath)
    #     print("结束！！！！！！！")

    """
    特殊情况一：有横向表格
    将相应的信息写入表格文件
    """
    def ExcelWrite_Xlsx_SpecialOne(self,iFilePath:str,iExcelClass,iTitleDict:dict,iPdmAllDict:dict,iIndexDict:dict
                       ,iSheetList:list,iAllZCName:dict,iAllZCNumber:dict,iAllLightNum:dict):
        print("开始写数据！！！！！！！")
        print("iTitleDict = ",iTitleDict)
        for i in iSheetList:
            sheet = iExcelClass.get_sheet_by_name(i)             # 读取sheet
            if iPdmAllDict[i] != []:
                for j in iPdmAllDict[i]:
                    z = re.split(r'@@',str(j))[1]
                    # 格式设定(尝试不知道能不能成功)
                    sheet.cell(iIndexDict[i][j], int(z)).alignment = Alignment(horizontal="center", vertical="center",wrapText=True)
                    sheet.cell(iIndexDict[i][j], int(z)+1).alignment = Alignment(horizontal="center", vertical="center",wrapText=True)
                    sheet.cell(iIndexDict[i][j], int(z)+2).alignment = Alignment(horizontal="center", vertical="center",wrapText=True)
                    # 数据写入
                    sheet.cell(iIndexDict[i][j], int(z)).value = iAllZCName[i][j]                # 主机厂
                    sheet.cell(iIndexDict[i][j], int(z)+1).value = iAllZCNumber[i][j]            # 灯具总成号
                    sheet.cell(iIndexDict[i][j], int(z)+2).value = iAllLightNum [i][j]           # 借用灯具数量
        print("结束！！！！！！！")
        # 文件保存
        print("保存！！！！！！！")
        # 文件保存
        iExcelClass.save(iFilePath) 
        print("结束！！！！！！！")

    """
    将相应的信息写入表格文件
    输入:
    文件存储地址、Excel文件对象、表头字典、PDM号字典、PDM行索引字典、Sheet列表
    字典详解：
    1.表头字典------->key:sheet名 value:对应表头的行列值列表(ITEM号所在行和列、主机厂所在行和列)
    2.PDM号字典------>key:sheet名 value:该sheet里的PDM号列表
    3.PDM行索引字典-->key:sheet名 value:数据字典-->key:PDM号 value:行索引
    """
    def ExcelWrite_Xlsx(self,iFilePath:str,iExcelClass,iTitleDict:dict,iPdmAllDict:dict,iIndexDict:dict
                       ,iSheetList:list,iAllZCName:dict,iAllZCNumber:dict,iAllLightNum:dict):
        print("开始写数据！！！！！！！")
        print("iTitleDict = ",iTitleDict)
        """
        引用实例化的表格类来写数据
        """
        for i in iSheetList:
            sheet = iExcelClass.get_sheet_by_name(i)             # 读取sheet
            if iPdmAllDict[i] != []:
                for j in iPdmAllDict[i]:
                    try:
                        # 格式设定(尝试不知道能不能成功)
                        sheet.cell(iIndexDict[i][j], iTitleDict[i][3]).alignment = Alignment(horizontal="center", vertical="center",wrapText=True)
                        sheet.cell(iIndexDict[i][j], iTitleDict[i][3]+1).alignment = Alignment(horizontal="center", vertical="center",wrapText=True)
                        sheet.cell(iIndexDict[i][j], iTitleDict[i][3]+2).alignment = Alignment(horizontal="center", vertical="center",wrapText=True)
                        # 数据写入
                        sheet.cell(iIndexDict[i][j], iTitleDict[i][3]).value = iAllZCName[i][j]                # 主机厂
                        sheet.cell(iIndexDict[i][j], iTitleDict[i][3]+1).value = iAllZCNumber[i][j]            # 灯具总成号
                        sheet.cell(iIndexDict[i][j], iTitleDict[i][3]+2).value = iAllLightNum [i][j]           # 借用灯具数量
                    except:
                        print("可能有合并单元格的问题在第{}行，第{}列".format(iIndexDict[i][j],iTitleDict[i][3]))
        print("结束！！！！！！！")
        # 文件保存
        print("保存！！！！！！！")
        # 文件保存
        iExcelClass.save(iFilePath) 
        print("结束！！！！！！！")

    """
    接收特殊表格信号
    """
    def SignalSpecial(self,iIsSpecial:int):
        self.IsSpecial = iIsSpecial

    """
    初始化
    """
    def Initial(self):
        # 网页操作变量
        self.options = None
        self.serv = None
        self.browser = None

        # Excel文件操作变量
        self.WorkBookRD = None              # xlrd 文件打开类
        self.WorkBookWT = None              # xlwt 文件写入类
        self.WorkBookOP = None              # openpyxl 文件打开写入类
        self.PDMIndexList:list = []         # PDM号在表格中的列位置
        self.TitleIndexList:int = 0         # 表头所在的行数
        self.NewMassageIndex:int = 0        # 填入数据的列位置  (n,n+1,n+2)

    """
    字典变量初始化
    在每次读取不同的Excel表格时进行初始化(内部初始化)
    """
    def InitialDic(self):
        self.TitleDict:dict = {}            # 存储表头行列信息的字典
        self.PdmAllDict:dict = {}           # 字典:key:sheet名称 value:PDM号列表
        self.IndexDict:dict = {}            # 字典:key:sheet名称 value:PDM号行号列表

    """
    是否为进行文件夹选择
    """
    def IsChooseFolder(self,iIsFolder:bool):
        self.IsFolder = iIsFolder

    def SetMassageClass(self,ip_MassageWindow:MassageWindow):
        self.m_MassageWindow = ip_MassageWindow

    """
    启动线程后的执行函数
    """
    def run(self):
        self.WrangMassageDic = {}
        # 文件夹操作
        if self.IsFolder:
            # 提取文件路径列表中的值来读取Excel文件
            for i in range(len(self.FilePathList)):
                AllZCName = {}   # 临时变量
                AllZCNumber = {} # 临时变量
                AllLightNum = {} # 临时变量
                # 判断是否含有特殊字符
                if len(re.split(r'@', str(self.FilePathList[i]), flags=re.IGNORECASE)) == 1:
                    if i < self.XlsNum:
                        self.InitialDic()       # 初始化字典变量
                        self.ExcelRead_Xlsx(self.FilePathList[i]+"x")   # 数据提取
                        sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                        for sheetName in self.SheetList:
                            pdmlist = []        # 临时变量:无数据的pdm号列表
                            for PDMNum in self.PdmAllDict[sheetName]:
                                self.SetPDMNumber(PDMNum)
                                # 无数据时的处理
                                if self.IsNone:
                                    pattern=re.compile("/")
                                    result = pattern.split(self.FilePathList[i])
                                    self.NoInformationFileName = result[-1]
                                    print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                    pdmlist.append(PDMNum)
                                    sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                            AllZCName[sheetName] = self.ZCNameDict
                            AllZCNumber[sheetName] = self.ZCNumberDict
                            AllLightNum[sheetName] = self.LightNumDict
                        # 有数据时写入Excel文件中
                        self.ExcelWrite_Xlsx(self.FilePathList[i],self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                            ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                            
                        self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中
                        # 结束
                    else:
                        self.InitialDic()       # 初始化字典变量
                        self.ExcelRead_Xlsx(self.FilePathList[i])
                        sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                        for sheetName in self.SheetList:
                            pdmlist = []        # 临时变量:无数据的pdm号列表
                            for PDMNum in self.PdmAllDict[sheetName]:
                                self.SetPDMNumber(PDMNum)
                                # 无数据时的处理
                                if self.IsNone:
                                    pattern=re.compile("/")
                                    result = pattern.split(self.FilePathList[i])
                                    self.NoInformationFileName = result[-1]
                                    print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                    pdmlist.append(PDMNum)
                                    sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                            AllZCName[sheetName] = self.ZCNameDict
                            AllZCNumber[sheetName] = self.ZCNumberDict
                            AllLightNum[sheetName] = self.LightNumDict
                        # 有数据时写入Excel文件中
                        self.ExcelWrite_Xlsx(self.FilePathList[i],self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                            ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                        
                        self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中
                        # 结束
                else:
                    if '总成' in re.split(r'@', str(self.FilePathList[i]), flags=re.IGNORECASE)[-1]:
                        if i < self.XlsNum:
                            self.InitialDic()       # 初始化字典变量
                            self.AssemblyItemExcel(self.FilePathList[i]+"x")   # 数据提取
                            sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                            for sheetName in self.SheetList:
                                pdmlist = []        # 临时变量:无数据的pdm号列表
                                for PDMNum in self.PdmAllDict[sheetName]:
                                    self.SetPDMNumber(PDMNum)
                                    # 无数据时的处理
                                    if self.IsNone:
                                        pattern=re.compile("/")
                                        result = pattern.split(self.FilePathList[i])
                                        self.NoInformationFileName = result[-1]
                                        print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                        pdmlist.append(PDMNum)
                                        sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                                AllZCName[sheetName] = self.ZCNameDict
                                AllZCNumber[sheetName] = self.ZCNumberDict
                                AllLightNum[sheetName] = self.LightNumDict
                            # 有数据时写入Excel文件中
                            self.ExcelWrite_Xlsx(self.FilePathList[i],self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                                ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                                
                            self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中
                            # 结束
                        else:
                            self.InitialDic()       # 初始化字典变量
                            self.AssemblyItemExcel(self.FilePathList[i])
                            sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                            for sheetName in self.SheetList:
                                pdmlist = []        # 临时变量:无数据的pdm号列表
                                for PDMNum in self.PdmAllDict[sheetName]:
                                    self.SetPDMNumber(PDMNum)
                                    # 无数据时的处理
                                    if self.IsNone:
                                        pattern=re.compile("/")
                                        result = pattern.split(self.FilePathList[i])
                                        self.NoInformationFileName = result[-1]
                                        print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                        pdmlist.append(PDMNum)
                                        sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                                AllZCName[sheetName] = self.ZCNameDict
                                AllZCNumber[sheetName] = self.ZCNumberDict
                                AllLightNum[sheetName] = self.LightNumDict
                            # 有数据时写入Excel文件中
                            self.ExcelWrite_Xlsx(self.FilePathList[i],self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                                ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                            
                            self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中
                            # 结束
                    elif '多表格' in re.split(r'@', str(self.FilePathList[i]), flags=re.IGNORECASE)[-1]:
                        if i < self.XlsNum:
                            self.InitialDic()       # 初始化字典变量
                            self.HorizontalArrangementExcel(self.FilePathList[i]+"x")   # 数据提取
                            sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                            for sheetName in self.SheetList:
                                pdmlist = []        # 临时变量:无数据的pdm号列表
                                for PDMNum in self.PdmAllDict[sheetName]:
                                    self.SetPDMNumber(PDMNum)
                                    # 无数据时的处理
                                    if self.IsNone:
                                        pattern=re.compile("/")
                                        result = pattern.split(self.FilePathList[i])
                                        self.NoInformationFileName = result[-1]
                                        print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                        pdmlist.append(PDMNum)
                                        sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                                AllZCName[sheetName] = self.ZCNameDict
                                AllZCNumber[sheetName] = self.ZCNumberDict
                                AllLightNum[sheetName] = self.LightNumDict
                            # 有数据时写入Excel文件中
                            self.ExcelWrite_Xlsx_SpecialOne(self.FilePathList[i],self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                                ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                                
                            self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中
                            # 结束
                        else:
                            self.InitialDic()       # 初始化字典变量
                            self.HorizontalArrangementExcel(self.FilePathList[i])
                            sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                            for sheetName in self.SheetList:
                                pdmlist = []        # 临时变量:无数据的pdm号列表
                                for PDMNum in self.PdmAllDict[sheetName]:
                                    self.SetPDMNumber(PDMNum)
                                    # 无数据时的处理
                                    if self.IsNone:
                                        pattern=re.compile("/")
                                        result = pattern.split(self.FilePathList[i])
                                        self.NoInformationFileName = result[-1]
                                        print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                        pdmlist.append(PDMNum)
                                        sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                                AllZCName[sheetName] = self.ZCNameDict
                                AllZCNumber[sheetName] = self.ZCNumberDict
                                AllLightNum[sheetName] = self.LightNumDict
                            # 有数据时写入Excel文件中
                            self.ExcelWrite_Xlsx_SpecialOne(self.FilePathList[i],self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                                ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                            
                            self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中
                            # 结束
                self.FinishSignal.emit(False) # 未执行结束发送False信号
            self.m_MassageWindow.SetData(self.WrangMassageDic)              # 将空值字典传入Massage类中
            print("SetPDMNumber(self.WrangMassageDic) = ",self.WrangMassageDic)
            self.FinishSignal.emit(True)      # 执行结束发送True信号

        # 单文件操作
        else:
            if len(re.split(r'@', str(self.FilePath), flags=re.IGNORECASE)) == 1:
                if self.FilePath[-1] == 's':
                    AllZCName = {}   # 临时变量
                    AllZCNumber = {} # 临时变量
                    AllLightNum = {} # 临时变量
                    self.InitialDic()       # 初始化字典变量
                    self.ExcelRead_Xlsx(self.FilePath+"x")   # 数据提取
                    sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                    for sheetName in self.SheetList:
                        pdmlist = []        # 临时变量:无数据的pdm号列表
                        for PDMNum in self.PdmAllDict[sheetName]:
                            self.SetPDMNumber(PDMNum)
                            # 无数据时的处理
                            if self.IsNone:
                                pattern=re.compile("/")
                                result = pattern.split(self.FilePath)
                                self.NoInformationFileName = result[-1]
                                print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                pdmlist.append(PDMNum)
                                sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                        AllZCName[sheetName] = self.ZCNameDict
                        AllZCNumber[sheetName] = self.ZCNumberDict
                        AllLightNum[sheetName] = self.LightNumDict
                    # 有数据时写入Excel文件中
                    self.ExcelWrite_Xlsx(self.FilePath,self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                            ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                    self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中
                    # 结束
                else:
                    AllZCName = {}   # 临时变量
                    AllZCNumber = {} # 临时变量
                    AllLightNum = {} # 临时变量
                    self.InitialDic()       # 初始化字典变量
                    self.ExcelRead_Xlsx(self.FilePath)
                    sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                    for sheetName in self.SheetList:
                        pdmlist = []        # 临时变量:无数据的pdm号列表
                        for PDMNum in self.PdmAllDict[sheetName]:
                            self.SetPDMNumber(PDMNum)
                            # 无数据时的处理
                            if self.IsNone:
                                pattern=re.compile("/")
                                result = pattern.split(self.FilePath)
                                self.NoInformationFileName = result[-1]
                                print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                pdmlist.append(PDMNum)
                                sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                        AllZCName[sheetName] = self.ZCNameDict
                        AllZCNumber[sheetName] = self.ZCNumberDict
                        AllLightNum[sheetName] = self.LightNumDict
                    # 有数据时写入Excel文件中
                    self.ExcelWrite_Xlsx(self.FilePath,self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                            ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                    self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中                
            else:
                if '多表格' in re.split(r'@', str(self.FilePath), flags=re.IGNORECASE)[-1]:
                    AllZCName = {}   # 临时变量
                    AllZCNumber = {} # 临时变量
                    AllLightNum = {} # 临时变量
                    self.InitialDic()       # 初始化字典变量
                    self.HorizontalArrangementExcel(self.FilePath) # 特殊情况一
                    sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                    for sheetName in self.SheetList:
                        pdmlist = []        # 临时变量:无数据的pdm号列表
                        for PDMNum in self.PdmAllDict[sheetName]:
                            self.SetPDMNumber(PDMNum)
                            # 无数据时的处理
                            if self.ZCName == "" or self.ZCNumber == "":
                                pattern=re.compile("/")
                                result = pattern.split(self.FilePath)
                                self.NoInformationFileName = result[-1]
                                print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                pdmlist.append(PDMNum)
                                sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                        AllZCName[sheetName] = self.ZCNameDict
                        AllZCNumber[sheetName] = self.ZCNumberDict
                        AllLightNum[sheetName] = self.LightNumDict
                    # 有数据时写入Excel文件中
                    self.ExcelWrite_Xlsx_SpecialOne(self.FilePath,self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                            ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                    self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中

                elif '总成' in re.split(r'@', str(self.FilePath), flags=re.IGNORECASE)[-1]:
                    AllZCName = {}   # 临时变量
                    AllZCNumber = {} # 临时变量
                    AllLightNum = {} # 临时变量
                    self.InitialDic()       # 初始化字典变量
                    self.AssemblyItemExcel(self.FilePath)
                    sheet_pdm = {}      # 临时变量:存储没有数据的sheet对应的pdm号信息
                    for sheetName in self.SheetList:
                        pdmlist = []        # 临时变量:无数据的pdm号列表
                        for PDMNum in self.PdmAllDict[sheetName]:
                            self.SetPDMNumber(PDMNum)
                            # 无数据时的处理
                            if self.ZCName == "" or self.ZCNumber == "":
                                pattern=re.compile("/")
                                result = pattern.split(self.FilePath)
                                self.NoInformationFileName = result[-1]
                                print("SetFileName(self.NoInformationFileName) = ",self.NoInformationFileName)
                                pdmlist.append(PDMNum)
                                sheet_pdm[sheetName] = pdmlist         # 将无数据的pdm号列表压入字典中
                        AllZCName[sheetName] = self.ZCNameDict
                        AllZCNumber[sheetName] = self.ZCNumberDict
                        AllLightNum[sheetName] = self.LightNumDict
                    # 有数据时写入Excel文件中
                    self.ExcelWrite_Xlsx(self.FilePath,self.WorkBookOP,self.TitleDict,self.PdmAllDict
                                            ,self.IndexDict,self.SheetList,AllZCName,AllZCNumber,AllLightNum)
                    self.WrangMassageDic[self.NoInformationFileName] = sheet_pdm   # 将字典存入错误信息字典中
        # 结束
        print("self.WrangMassageDic = ",self.WrangMassageDic)
        self.m_MassageWindow.SetData(self.WrangMassageDic)              # 将空值字典传入Massage类中
        self.FinishSignal.emit(True)      # 执行结束发送True信号
        self.IsSpecial = 0
