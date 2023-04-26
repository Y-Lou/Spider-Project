# 界面类
from MassageDialog import Ui_MassageDialog

# xlsx文件写入并保存
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment

# PyQt5类
from PyQt5 import QtCore
from PyQt5.QtWidgets import QDialog,QFileDialog
from PyQt5.QtWidgets  import  QHeaderView,QTableWidgetItem,QTreeWidgetItem

class MassageWindow(QDialog):
    def __init__(self) -> None: # 构造函数
        super().__init__()
        self.ui:Ui_MassageDialog = Ui_MassageDialog()
        self.ui.setupUi(self)

        # 成员变量
        self.oMassageDic:dict = {"Folder":"","Excel_File":{}}  # 信息收集字典->(文件夹名、Excel文件名、Excel->PDM(字典))
        self.FolderName:str = ""                                              # 打开文件夹名称
        self.FileName:str = ""                                                # Excel文件名称
        self.oFilePath:str = ""                                               # 文件保存地址
        self.RowConut:int = 0                                                 # 总共多少PDM号没有数据

        # Excel文件操作变量
        self.WorkBookOP = None              # openpyxl 文件打开写入类

        # 信号连接
        self.SignalConnect()

    """
    信号连接
    """
    def SignalConnect(self):
        self.ui.pushButton.clicked.connect(self.MassageExcelWrite)   # 连接文件保存函数
        self.ui.tabWidget.currentChanged.connect(self.ShowMassage)   # 连接标签页窗口切换信号
  
    """
    获取文件夹名称
    """
    def SetFolderName(self,iFoldName:str):
        self.FolderName =iFoldName                         # 获取文件夹名
        self.oMassageDic["Folder"] = self.FolderName       # 压入字典

    """
    获取Excel文件名称
    值更新:Excel文件名字典->key:Excel_File value:Excel文件名列表
    """
    def SetData(self,iFileName:dict):
        self.oMassageDic["Excel_File"] = iFileName    # 压入字典

    """
    获取无数据的PDM编号
    值更新:总数据字典>key:PDMNumber value:无数据的PDM号字典->key:Excel文件名 value:sheet名->PDM号字典
    """
    def SetPDMNumber(self,iPDMNum:dict):
        self.oMassageDic["PDMNumber"] = iPDMNum                 # 写入字典
        self.RowConut = 0                                       # 记录列表窗口行数
        for i in self.oMassageDic["PDMNumber"].keys():          # 获取文件名
            for j in self.oMassageDic["PDMNumber"][i].keys():   # 获取sheet名
                self.RowConut += len(self.oMassageDic["PDMNumber"][i][j])
    """
    将数据填入Excel文件
    """
    def MassageExcelWrite(self):
        self.WorkBookOP = Workbook()
        # 激活 worksheet
        ws = self.WorkBookOP.active
        ws.title = "空数据信息"
        # 设置每列的标题内容
        title = ['文件夹名','文件名','sheet','PDM编号']
        # 将标题写入表格
        for i,d in enumerate(title):
            ws.cell(1,i+1).value = d
         # 设置每列的宽度    
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 20
        # 设置内容格式水平居中垂直居中
        ws[f'A{1}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'B{1}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'C{1}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'D{1}'].alignment = Alignment(horizontal="center", vertical="center")
        # 数据写入
        TitleCont = 2 # 临时变量:表格的行索引
        SheetCont = 2 # 临时变量:sheet名行索引
        PDMCont = 2 # 临时变量:因为要合并单元格,单独的评论列索引
        HB = [2] # 临时变量:记录表格中的列索引变化,用于合并单元格(文件名)
        HS = [2] # 临时变量:记录表格中的列索引变化,用于合并单元格(sheet)
        ws[f'A{TitleCont}'].alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(TitleCont,1).value = self.oMassageDic["Folder"]
        for i in self.oMassageDic["Excel_File"].keys():
            ws[f'B{TitleCont}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{TitleCont}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{TitleCont}'].alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(TitleCont,2).value = i
            for j in self.oMassageDic["Excel_File"][i].keys():
                ws.cell(SheetCont,3).value = j
                for k in self.oMassageDic["Excel_File"][i][j]:
                    ws.cell(PDMCont,4).value = k
                    PDMCont += 1
                SheetCont = PDMCont
                HS.append(SheetCont)
            TitleCont = PDMCont
            HB.append(TitleCont)
        if len(HB) != 1 and len(HS) != 1:
            # 合并单元格
            ws.merge_cells(f'A{2}:A{HB[-1]-1}')
            for hb in range(len(HB)-1):
                ws.merge_cells(f'B{HB[hb]}:B{HB[hb+1]-1}')
            for hs in range(len(HS)-1):
                ws.merge_cells(f'C{HS[hs]}:C{HS[hs+1]-1}')
            # 文件保存
            self.oFilePath, ok = QFileDialog.getSaveFileName(self,'错误数据保存','./','(*.xlsx)')
            if ok:
                self.WorkBookOP.save(self.oFilePath)

    def Initial(self):
        # 成员变量
        self.oMassageDic:dict = {"Folder":"","Excel_File":[],"PDMNumber":{}}  # 信息收集字典->(文件夹名、Excel文件名、Excel->PDM(字典))
        self.FolderName:str = ""                                              # 打开文件夹名称
        self.FileName:str = ""                                                # Excel文件名称
        self.oFilePath:str = ""                                               # 文件保存地址

        # Excel文件操作变量
        self.WorkBookOP = None              # openpyxl 文件打开写入类

    """
    数据显示
    """
    def ShowMassage(self):
        if self.ui.tabWidget.currentIndex() == 0:
            self.TreeWidget()

    """
    起始数据显示
    显示表格窗口
    """
    def OriginShowMassage(self):
        # self.TableWidget()
        self.TreeWidget()
    
    """
    表格窗口
    """
    # def TableWidget(self):
    #     self.ui.MassageTable.clear()        # 表格清空
    #     self.ui.MassageTable.setColumnCount(4) # 设置为4列
    #     self.ui.MassageTable.setRowCount(self.RowConut+1) # 设置行数
    #     self.ui.MassageTable.setHorizontalHeaderLabels(["文件夹名","文件名","sheet","PDM号"])  # 设置表头
    #     self.ui.MassageTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)     # 宽度适应
    #     self.ui.MassageTable.horizontalHeader().setStyleSheet("QHeaderView::section{background:lightgreen;}") # 设置表头颜色
    #     item4rowIndex:int = 0
    #     item1 =  QTableWidgetItem(self.oMassageDic["Folder"]) # 文件夹名
    #     self.ui.MassageTable.setItem( item4rowIndex, 0, item1)   # 第一列信息文件夹名
    #     for i in self.oMassageDic["Excel_File"].keys():
    #         item2 =  QTableWidgetItem(i)
    #         self.ui.MassageTable.setItem( item4rowIndex, 1, item2)   # 第二列信息文件名
    #         for j in self.oMassageDic["Excel_File"][i].keys():
    #             item3 =  QTableWidgetItem(j)
    #             self.ui.MassageTable.setItem( item4rowIndex, 2, item3)   # 第三列信息sheet名
    #             for k in self.oMassageDic["Excel_File"][i][j]:
    #                 item4 =  QTableWidgetItem(k)
    #                 self.ui.MassageTable.setItem( item4rowIndex, 3, item4)   # 第四列信息PDM号
    #                 item4rowIndex+=1
    #     # 表格显示
    #     self.ui.MassageTable.show()

    """
    列表窗口
    """
    # def ListWidget(self):
    #     # 信息清空
    #     self.ui.FileList.clear()
    #     self.ui.ExcelList.clear()
    #     self.ui.PDMList.clear()
    #     # 信息填入
    #     self.ui.FileList.addItems(list(self.oMassageDic["Folder"]))
    #     self.ui.ExcelList.addItems(list(self.oMassageDic["Excel_File"]))
    #     index = self.oMassageDic["Excel_File"][0]
    #     sheetIndex = list(self.oMassageDic["PDMNumber"][index].keys())[0]
    #     self.ui.PDMList.addItems(list(self.oMassageDic["PDMNumber"][index][sheetIndex]))
    #     # 列表显示
    #     self.ui.FileList.show()
    #     self.ui.ExcelList.show()
    #     self.ui.PDMList.show()
    
    """
    列表窗口的槽函数
    """
    # def ListWidgetSlot(self,item):
    #     self.ui.PDMList.clear()
    #     sheetIndex = self.self.oMassageDic["PDMNumber"][item.text()].keys()
    #     for i in sheetIndex:
    #         self.ui.FileList.addItems(list(self.oMassageDic["PDMNumber"][item.text()][i]))  # 填入无数据的PDM号
    #     # 列表显示
    #     self.ui.PDMList.show()

    """
    树形窗口
    """
    def TreeWidget(self):
        # 清空树结构
        self.ui.MassageTree.clear()
        # 设置列数
        self.ui.MassageTree.setColumnCount(1)
        # 设置表头
        self.ui.MassageTree.setHeaderLabels(['空数据信息'])
        # 设置根节点          文件夹名
        root= QTreeWidgetItem(self.ui.MassageTree)
        root.setText(0,self.oMassageDic["Folder"])           # 填入根节点信息
        # 设置第一级枝节点    文件名
        for i in self.oMassageDic["Excel_File"].keys():           # 获取文件名
            child1 = QTreeWidgetItem(root)                 # 设置第一级子节点
            child1.setText(0,i)                              # 填入第一级子节点信息
            # 设置第二级枝节点    sheet名
            for j in self.oMassageDic["Excel_File"][i].keys():   # 获取对应文件中的sheet名
                child2 = QTreeWidgetItem(child1)           # 设置二级子节点
                child2.setText(0,j)                          # 填入第二级子节点信息
                for k in self.oMassageDic["Excel_File"][i][j]:   # 获取对应sheet中的无数据PDM号
                    child3 = QTreeWidgetItem(child2)       # 设置三级子节点
                    child3.setText(0,str(k))                 # 填入第三级子节点信息
        self.ui.MassageTree.show()
